import openpyxl

def openWorkbookSheet(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet

def getRowCount(sheet):
    return sheet.max_row

def getColumnCount(sheet):
    return sheet.max_column

def readData(sheet, rowno, columnno):
    return sheet.cell(row=rowno, column=columnno).value

def writeData(file, sheetName, rowno, columnno, inpData):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowno, column=columnno).value = inpData
    workbook.save(file)
